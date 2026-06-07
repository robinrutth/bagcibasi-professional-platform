from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .calculations import calculate_operation
from .models import Customer, Shipment
from .repository import to_shipment_model_payload


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("ı", "i")


def as_float(value: Any, default: float = 0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    return float(text)


def as_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(value))
    if value:
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(str(value), fmt).date()
            except ValueError:
                pass
    return date.today()


def split_route(route: str) -> tuple[str, str]:
    parts = [part.strip() for part in route.replace("–", "-").split("-") if part.strip()]
    if len(parts) >= 2:
        return parts[0], parts[1]
    return "Manisa", "İstanbul"


def find_header_indexes(header_row: tuple[Any, ...]) -> dict[str, int]:
    headers = {normalize(value): index for index, value in enumerate(header_row)}
    aliases = {
        "date": ["tarih"],
        "customer": ["müşteri", "musteri"],
        "route": ["güzergah", "guzergah"],
        "distance": ["mesafe (km)", "mesafe"],
        "vehicle": ["araç tipi", "arac tipi"],
        "co2": ["co2 (kg)", "co2"],
        "cost": ["alış maliyeti", "alis maliyeti"],
        "invoice": ["kesilen fatura", "fatura"],
        "profit": ["brüt kâr", "brut kar", "kar"],
        "margin": ["kâr marjı", "kar marji"],
        "payment": ["ödeme durumu", "odeme durumu"],
        "status": ["operasyon durumu", "durum"],
    }
    found: dict[str, int] = {}
    for key, candidates in aliases.items():
        for candidate in candidates:
            if candidate in headers:
                found[key] = headers[candidate]
                break
    return found


def get_cell(row: tuple[Any, ...], indexes: dict[str, int], key: str, default: Any = None) -> Any:
    index = indexes.get(key)
    if index is None or index >= len(row):
        return default
    return row[index]


def import_operations_excel(db: Session, content: bytes) -> dict:
    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook["OPERASYON MERKEZI"] if "OPERASYON MERKEZI" in workbook.sheetnames else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": ["Excel boş görünüyor."]}

    indexes = find_header_indexes(rows[0])
    imported = 0
    skipped = 0
    errors: list[str] = []

    for row_number, row in enumerate(rows[1:], start=2):
        try:
            customer_name = str(get_cell(row, indexes, "customer", "") or "").strip()
            route = str(get_cell(row, indexes, "route", "") or "").strip()
            if not customer_name or not route:
                skipped += 1
                continue

            origin, destination = split_route(route)
            delivery_date = as_date(get_cell(row, indexes, "date"))
            distance = as_float(get_cell(row, indexes, "distance"), 0)
            vehicle = str(get_cell(row, indexes, "vehicle", "") or "").strip() or "Kamyon"
            cost = as_float(get_cell(row, indexes, "cost"), 0)
            invoice = as_float(get_cell(row, indexes, "invoice"), 0)
            profit = as_float(get_cell(row, indexes, "profit"), invoice - cost if invoice and cost else 0)
            margin = as_float(get_cell(row, indexes, "margin"), profit / invoice if invoice else 0)
            co2 = as_float(get_cell(row, indexes, "co2"), 0)
            status = str(get_cell(row, indexes, "status", "Hazırlanıyor") or "Hazırlanıyor").strip()

            exists = db.scalar(
                select(Shipment).where(
                    Shipment.customer_name == customer_name,
                    Shipment.origin == origin,
                    Shipment.destination == destination,
                    Shipment.delivery_date == delivery_date,
                )
            )
            customer = db.scalar(select(Customer).where(Customer.name == customer_name))
            if not customer:
                customer = Customer(name=customer_name, sector="Excel Import", payment_terms="Vadeli", risk_level="Düşük")
                db.add(customer)
                db.flush()

            calculated = calculate_operation(origin, destination, "Genel Yük", 10, delivery_date)
            model_payload = to_shipment_model_payload(calculated)
            model_payload.update(
                {
                    "distance_km": distance or model_payload["distance_km"],
                    "vehicle_type": vehicle,
                    "cost_amount": cost or model_payload["cost_amount"],
                    "invoice_amount": invoice or model_payload["invoice_amount"],
                    "profit_amount": profit or model_payload["profit_amount"],
                    "profit_margin": margin or model_payload["profit_margin"],
                    "co2_kg": co2 or model_payload["co2_kg"],
                    "risk_level": "Orta" if status.lower() == "yolda" else "Düşük",
                }
            )
            if exists:
                exists.customer_id = customer.id
                exists.status = status
                for key, value in model_payload.items():
                    setattr(exists, key, value)
                skipped += 1
                continue

            db.add(
                Shipment(
                    customer_id=customer.id,
                    customer_name=customer_name,
                    status=status,
                    **model_payload,
                )
            )
            imported += 1
        except Exception as exc:
            errors.append(f"Satır {row_number}: {exc}")

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors[:10]}
