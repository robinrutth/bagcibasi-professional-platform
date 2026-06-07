from __future__ import annotations

import csv
from collections.abc import Iterable, Iterator
from datetime import date
from io import BytesIO, StringIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Customer, Shipment


IMPORT_CUSTOMER_COLUMNS = ["ad", "email", "telefon", "adres", "sehir", "vergi_no"]
IMPORT_SHIPMENT_COLUMNS = ["musteri_email", "origin", "destination", "vehicle_type", "weight_kg", "distance_km", "status"]

SHIPMENT_COLUMNS = [
    "id",
    "müşteri",
    "origin",
    "destination",
    "vehicle_type",
    "weight_kg",
    "distance_km",
    "carbon_emission",
    "status",
    "created_at",
]

CUSTOMER_COLUMNS = ["id", "ad", "email", "telefon", "şehir", "toplam sevkiyat", "toplam CO2"]


def _safe_value(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _shipment_row(shipment: Shipment) -> dict[str, Any]:
    return {
        "id": str(shipment.id),
        "müşteri": shipment.customer.name if shipment.customer else shipment.customer_name,
        "origin": shipment.origin,
        "destination": shipment.destination,
        "vehicle_type": shipment.vehicle_type,
        "weight_kg": shipment.weight_kg,
        "distance_km": shipment.distance_km,
        "carbon_emission": shipment.co2_kg,
        "status": shipment.status,
        "created_at": shipment.created_at,
    }


def _customer_row(customer: Customer) -> dict[str, Any]:
    active_shipments = [shipment for shipment in customer.shipments if not shipment.is_deleted]
    return {
        "id": str(customer.id),
        "ad": customer.name,
        "email": customer.email,
        "telefon": customer.phone,
        "şehir": customer.city,
        "toplam sevkiyat": len(active_shipments),
        "toplam CO2": round(sum(shipment.co2_kg or 0 for shipment in active_shipments), 4),
    }


def _csv_lines(rows: Iterable[dict[str, Any]], columns: list[str]) -> Iterator[str]:
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)

    for row in rows:
        writer.writerow({column: _safe_value(row.get(column)) for column in columns})
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)


def export_shipments_csv(shipments: Iterable[Shipment]) -> Iterator[str]:
    return _csv_lines((_shipment_row(shipment) for shipment in shipments), SHIPMENT_COLUMNS)


def export_customers_csv(customers: Iterable[Customer]) -> Iterator[str]:
    return _csv_lines((_customer_row(customer) for customer in customers), CUSTOMER_COLUMNS)


def _append_rows(sheet, columns: list[str], rows: Iterable[dict[str, Any]]) -> None:
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="16A34A")
    for row in rows:
        sheet.append([_safe_value(row.get(column)) for column in columns])
    _auto_width(sheet)


def _auto_width(sheet) -> None:
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 44)


def _workbook_bytes(workbook: Workbook) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _read_rows(file: bytes | BytesIO, filename: str | None = None) -> list[dict[str, Any]]:
    content = file if isinstance(file, bytes) else file.getvalue()
    if not content:
        raise ValueError("Bos dosya yuklenemez")
    suffix = (filename or "").lower()
    if suffix.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV dosyasi okunamadi") from exc
        return list(csv.DictReader(StringIO(text)))
    try:
        workbook = pd.read_excel(BytesIO(content), dtype=object)
    except Exception as exc:
        raise ValueError("Excel dosyasi okunamadi") from exc
    return workbook.fillna("").to_dict("records")


def _clean_row(row: dict[str, Any]) -> dict[str, Any]:
    return {str(key).strip(): (value.strip() if isinstance(value, str) else value) for key, value in row.items()}


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def validate_customer_row(row: dict[str, Any]) -> dict[str, Any]:
    row = _clean_row(row)
    missing = [column for column in IMPORT_CUSTOMER_COLUMNS if column not in row]
    if missing:
        raise ValueError(f"Eksik kolon: {', '.join(missing)}")
    if _blank(row.get("ad")):
        raise ValueError("ad zorunlu")
    if _blank(row.get("email")):
        raise ValueError("email zorunlu")
    return {
        "name": str(row["ad"]),
        "email": str(row["email"]).lower(),
        "phone": None if _blank(row.get("telefon")) else str(row["telefon"]),
        "address": None if _blank(row.get("adres")) else str(row["adres"]),
        "city": None if _blank(row.get("sehir")) else str(row["sehir"]),
        "tax_number": None if _blank(row.get("vergi_no")) else str(row["vergi_no"]),
        "sector": None,
        "payment_terms": "Vadeli",
        "risk_level": "Dusuk",
        "notes": None,
        "is_active": True,
    }


def validate_shipment_row(row: dict[str, Any]) -> dict[str, Any]:
    row = _clean_row(row)
    missing = [column for column in IMPORT_SHIPMENT_COLUMNS if column not in row]
    if missing:
        raise ValueError(f"Eksik kolon: {', '.join(missing)}")
    for column in ("musteri_email", "origin", "destination", "vehicle_type", "weight_kg", "distance_km"):
        if _blank(row.get(column)):
            raise ValueError(f"{column} zorunlu")
    try:
        weight_kg = float(row["weight_kg"])
        distance_km = float(row["distance_km"])
    except (TypeError, ValueError) as exc:
        raise ValueError("weight_kg ve distance_km sayisal olmali") from exc
    if weight_kg <= 0 or distance_km < 0:
        raise ValueError("weight_kg pozitif, distance_km sifir veya pozitif olmali")
    tonnage = max(weight_kg / 1000, 0.1)
    if tonnage > 26:
        raise ValueError("weight_kg 26000 degerinden buyuk olamaz")
    return {
        "customer_email": str(row["musteri_email"]).lower(),
        "origin": str(row["origin"]),
        "destination": str(row["destination"]),
        "vehicle_type": str(row["vehicle_type"]),
        "weight_kg": weight_kg,
        "distance_km": distance_km,
        "status": str(row.get("status") or "Hazirlaniyor"),
        "cargo_type": "Genel Yuk",
        "tonnage": tonnage,
        "delivery_date": date.today(),
    }


def parse_customer_excel(file: bytes | BytesIO, filename: str | None = None) -> list[dict[str, Any]]:
    return [row for _, row, error in generate_import_rows(file, filename, validate_customer_row) if not error]


def parse_shipment_excel(file: bytes | BytesIO, filename: str | None = None) -> list[dict[str, Any]]:
    return [row for _, row, error in generate_import_rows(file, filename, validate_shipment_row) if not error]


def generate_import_rows(
    file: bytes | BytesIO,
    filename: str | None,
    validator,
) -> list[tuple[int, dict[str, Any], str | None]]:
    rows = _read_rows(file, filename)
    if not rows:
        raise ValueError("Dosyada import edilecek satir bulunamadi")
    result = []
    for index, row in enumerate(rows, start=2):
        try:
            result.append((index, validator(row), None))
        except ValueError as exc:
            result.append((index, {}, str(exc)))
    return result


def generate_import_template(type: str) -> BytesIO:
    columns = IMPORT_CUSTOMER_COLUMNS if type == "customers" else IMPORT_SHIPMENT_COLUMNS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import"
    _append_rows(sheet, columns, [])
    return _workbook_bytes(workbook)


def export_shipments_excel(shipments: Iterable[Shipment]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sevkiyatlar"
    _append_rows(sheet, SHIPMENT_COLUMNS, (_shipment_row(shipment) for shipment in shipments))
    return _workbook_bytes(workbook)


def export_carbon_report_excel(summary: dict[str, Any], period: str) -> BytesIO:
    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="16A34A")

    def style_header(row) -> None:
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill

    def rounded(value: Any) -> float:
        return round(float(value or 0), 4)

    top_routes = summary.get("top_routes", []) or []
    top_route = ""
    if top_routes:
        top_route = f"{top_routes[0].get('origin', '')} → {top_routes[0].get('destination', '')}"

    overview = workbook.active
    overview.title = "Özet"
    overview.append(["metrik", "değer"])
    overview.append(["Dönem", period])
    overview.append(["Toplam CO2 (kg)", rounded(summary.get("total_co2", 0))])
    overview.append(["Toplam Sevkiyat", summary.get("total_shipments", 0)])
    overview.append(["Ortalama CO2/Sevkiyat (kg)", rounded(summary.get("avg_co2", 0))])
    overview.append(["En Yüksek CO2'lu Rota", top_route])
    overview.append(["ISO 14083:2023 Uyumlu", "Evet"])
    overview.append(["Emisyon Metodolojisi", "WTW (Well-to-Wheel)"])
    overview.append(["Referans Standart", "EMEP/EEA 2023 / IPCC AR6"])
    style_header(overview[1])
    _auto_width(overview)

    shipment_sheet = workbook.create_sheet("Sevkiyat Detayları")
    shipment_columns = [
        "Müşteri",
        "Çıkış",
        "Varış",
        "Araç Tipi",
        "Ağırlık (kg)",
        "Mesafe (km)",
        "CO2 (kg)",
        "CO2 Verimi (gCO2e/ton-km)",
        "Fatura (₺)",
        "Durum",
        "Tarih",
    ]
    shipment_sheet.append(shipment_columns)
    style_header(shipment_sheet[1])
    for shipment in summary.get("shipments", []) or []:
        weight_kg = float(shipment.get("weight_kg", 0) or 0)
        distance_km = float(shipment.get("distance_km", 0) or 0)
        co2_kg = float(shipment.get("co2_kg", 0) or 0)
        weight_ton = max(weight_kg / 1000, 0.001)
        efficiency = round((co2_kg * 1000) / (weight_ton * distance_km), 4) if distance_km > 0 else ""
        shipment_sheet.append(
            [
                shipment.get("musteri", ""),
                shipment.get("origin", ""),
                shipment.get("destination", ""),
                shipment.get("vehicle_type", ""),
                rounded(weight_kg),
                rounded(distance_km),
                rounded(co2_kg),
                efficiency,
                rounded(shipment.get("invoice_amount", 0)),
                shipment.get("status", ""),
                _safe_value(shipment.get("created_at", "")),
            ]
        )
    _auto_width(shipment_sheet)

    vehicle_sheet = workbook.create_sheet("Araç Tipi Dağılımı")
    vehicle_sheet.append(["Araç Tipi", "Toplam CO2 (kg)", "Oran (%)"])
    style_header(vehicle_sheet[1])
    by_vehicle = summary.get("by_vehicle", []) or []
    total_vehicle_co2 = sum(float(row.get("co2", 0) or 0) for row in by_vehicle)
    for row in by_vehicle:
        co2 = float(row.get("co2", 0) or 0)
        ratio = round((co2 / total_vehicle_co2) * 100, 2) if total_vehicle_co2 > 0 else 0
        vehicle_sheet.append([row.get("vehicle_type", ""), rounded(co2), ratio])
    _auto_width(vehicle_sheet)

    methodology_sheet = workbook.create_sheet("ISO 14083 Metodoloji")
    methodology_sheet.append(["ISO 14083:2023 Emisyon Metodoloji Beyanı"])
    methodology_sheet["A1"].font = Font(bold=True, size=13)
    methodology_sheet.append([])
    methodology_sheet.append(["Parametre", "Değer"])
    style_header(methodology_sheet[3])
    methodology_sheet.append(["Standart", "ISO 14083:2023 – Uluslararası Karbondioksit Emisyon Raporlama"])
    methodology_sheet.append(["Kapsam", "WTW – Well-to-Wheel (Kuyudan Tekere)"])
    methodology_sheet.append(["Emisyon Faktörü Kaynağı", "EMEP/EEA Hava Kirleticiler Envanter Rehberi 2023"])
    methodology_sheet.append(["Küresel Isınma Potansiyeli", "IPCC AR6 – CO2e dönüşümü"])
    methodology_sheet.append(["Hesaplama Birimi", "gCO2e/ton-km (verimlilik) ve kg CO2e (mutlak)"])
    methodology_sheet.append(["Araç Kategorileri", "Kamyon, TIR, Minivan, Van, Motokurye"])
    methodology_sheet.append(["Yakıt Tipleri", "Dizel, Benzin, Elektrik, Biyodizel"])
    methodology_sheet.append(["Euro Normları", "Euro 3, Euro 4, Euro 5, Euro 6"])
    methodology_sheet.append(["Rapor Üreticisi", "Vorxa Logistics AI Platform"])
    methodology_sheet.append(["Beyan Tarihi", date.today().isoformat()])
    _auto_width(methodology_sheet)

    return _workbook_bytes(workbook)
