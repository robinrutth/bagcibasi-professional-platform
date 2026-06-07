from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.services.export_service import (
    CUSTOMER_COLUMNS,
    IMPORT_CUSTOMER_COLUMNS,
    IMPORT_SHIPMENT_COLUMNS,
    SHIPMENT_COLUMNS,
    export_carbon_report_excel,
    export_customers_csv,
    export_shipments_csv,
    export_shipments_excel,
    parse_customer_excel,
    parse_shipment_excel,
)


def _csv_rows(chunks):
    content = "".join(chunks)
    return list(csv.DictReader(StringIO(content)))


def _xlsx_bytes(columns, rows):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(columns)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def test_shipments_csv_contains_expected_columns(sample_shipments):
    rows = _csv_rows(export_shipments_csv(sample_shipments))
    assert rows
    assert list(rows[0].keys()) == SHIPMENT_COLUMNS
    assert rows[0]["müşteri"] == "Api Customer"
    assert float(rows[0]["carbon_emission"]) == 42


def test_shipments_excel_has_formatted_header(sample_shipments):
    workbook = load_workbook(export_shipments_excel(sample_shipments))
    sheet = workbook["Sevkiyatlar"]
    assert [cell.value for cell in sheet[1]] == SHIPMENT_COLUMNS
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].fill.fgColor.rgb == "0016A34A"
    assert sheet.column_dimensions["A"].width >= 12


def test_export_endpoints_apply_filters(client, auth_headers, sample_customer, sample_shipments):
    response = client.get(
        f"/api/v1/exports/shipments/csv?start_date=2026-05-01&end_date=2026-05-31&status=Yolda&vehicle_type=truck&customer_id={sample_customer.id}",
        headers=auth_headers("viewer"),
    )
    assert response.status_code == 200
    rows = list(csv.DictReader(StringIO(response.text)))
    assert len(rows) == 1
    assert rows[0]["status"] == "Yolda"
    assert response.headers["content-disposition"].startswith('attachment; filename="shipments_')


def test_empty_exports_do_not_error(client, auth_headers, db_session):
    response = client.get("/api/v1/exports/shipments/csv?start_date=2030-01-01&end_date=2030-01-31", headers=auth_headers("admin"))
    assert response.status_code == 200
    assert response.text.strip() == ",".join(SHIPMENT_COLUMNS)

    workbook = load_workbook(BytesIO(client.get("/api/v1/exports/shipments/excel?start_date=2030-01-01", headers=auth_headers("admin")).content))
    assert workbook["Sevkiyatlar"].max_row == 1


def test_carbon_and_customer_exports(sample_customer, sample_shipments):
    summary = {
        "total_co2": 74,
        "by_vehicle": [{"vehicle_type": "truck", "co2": 74}],
        "top_routes": [{"origin": "Manisa", "destination": "Ankara", "vehicle_type": "truck", "co2": 42, "shipment_count": 1}],
    }
    workbook = load_workbook(export_carbon_report_excel(summary, "monthly"))
    assert workbook.sheetnames == ["Özet", "Araç Tipi Dağılımı", "Rota Bazlı Detay"]
    assert workbook["Özet"]["B2"].value == 74

    rows = _csv_rows(export_customers_csv([sample_customer]))
    assert list(rows[0].keys()) == CUSTOMER_COLUMNS
    assert rows[0]["toplam sevkiyat"] == "3"
    assert float(rows[0]["toplam CO2"]) == 74


def test_parse_customer_excel_returns_valid_rows():
    content = _xlsx_bytes(
        IMPORT_CUSTOMER_COLUMNS,
        [["Acme Lojistik", "acme@example.com", "555", "Adres", "Manisa", "123"]],
    )
    rows = parse_customer_excel(content, "customers.xlsx")
    assert rows == [
        {
            "name": "Acme Lojistik",
            "email": "acme@example.com",
            "phone": "555",
            "address": "Adres",
            "city": "Manisa",
            "tax_number": "123",
            "sector": None,
            "payment_terms": "Vadeli",
            "risk_level": "Dusuk",
            "notes": None,
            "is_active": True,
        }
    ]


def test_parse_shipment_excel_skips_invalid_rows():
    content = _xlsx_bytes(
        IMPORT_SHIPMENT_COLUMNS,
        [
            ["acme@example.com", "Manisa", "Ankara", "truck", 12000, 450, "Yolda"],
            ["", "Manisa", "Ankara", "truck", 12000, 450, "Yolda"],
        ],
    )
    rows = parse_shipment_excel(content, "shipments.xlsx")
    assert len(rows) == 1
    assert rows[0]["customer_email"] == "acme@example.com"
    assert rows[0]["weight_kg"] == 12000


def test_import_customers_updates_duplicate_email(client, auth_headers, sample_customer, db_session):
    content = _xlsx_bytes(
        IMPORT_CUSTOMER_COLUMNS,
        [["Updated Customer", sample_customer.email, "999", "Yeni adres", "Izmir", "999"]],
    )
    response = client.post(
        "/api/v1/imports/customers",
        headers=auth_headers("admin"),
        files={"file": ("customers.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    assert response.json()["success"] == 1
    db_session.refresh(sample_customer)
    assert sample_customer.name == "Updated Customer"
    assert sample_customer.city == "Izmir"


def test_parse_empty_import_file_errors():
    try:
        parse_customer_excel(b"", "customers.xlsx")
    except ValueError as exc:
        assert "Bos dosya" in str(exc)
    else:
        raise AssertionError("empty import file should fail")
